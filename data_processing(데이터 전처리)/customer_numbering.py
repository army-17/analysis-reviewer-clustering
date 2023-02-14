# ------------------- 코드 파일 설명 -------------------#
'''
- 리뷰어 키 부여하는 코드 파일
- 11번째 컬럼에 부여함(맨뒤에 추가했다는 뜻)
- customer_data1(125) 폴더안 125개 파일 읽어와서
  customer_data1_numbering(125) 폴더안에 키컬럼 추가된 125개 파일 생성

'''

import os
from openpyxl import load_workbook


# ------------------- main -------------------#
file_list = os.listdir("C:\\Users\\mingyeongkim\\Desktop\\clustering_project(1250)\\data\\customer_data(1250)")

print(file_list)
print(len(file_list))

num = 1
for file_name in file_list:

    direct = "C:\\Users\\mingyeongkim\\Desktop\\clustering_project(1250)\\data\\customer_data(1250)\\{0}".format(file_name)

    wb = load_workbook(direct)
    ws = wb.active

    ws['L1'] ='Customer_Key'
    ws['L2'] = num

    write_direct = "C:\\Users\\mingyeongkim\\Desktop\\clustering_project(1250)\\data\\customer_data_numbering(1250)\\{0}".format(file_name)
    wb.save(write_direct)

    num = num + 1
